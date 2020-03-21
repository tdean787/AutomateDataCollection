import openpyxl
import os
import glob
import pandas
from openpyxl import load_workbook
from openpyxl import Workbook
import selenium
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options 
import tkinter
from tkinter import filedialog

# driver = webdriver.Firefox()
# workbook = load_workbook('address_search.xlsx')
donortestinfo = pandas.read_excel('testsheet.xlsx')
donors = donortestinfo.to_dict('records')
print(donors)

dict_cities = []
for donor in donors:
    # print(donor["Name"])
    print(donor["City"])
    dict_cities.append(donor["City"])

print(dict_cities)