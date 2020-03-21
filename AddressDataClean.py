import openpyxl
import os
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
import time

workbook = load_workbook('updated_info.xlsx')
sheets = workbook.sheetnames
ws1 = workbook[sheets[0]]

names = []
for row in ws1.iter_rows(min_row=1,max_row=None, min_col=2, max_col =2):
    for cell in row:
        if "Helotes" in cell.value:
            cell.offset(0,1).value = "Helotes"
        elif "Shavano Park" in cell.value:
            cell.offset(0,1).value = "Shavano Park"
        elif "Hollywood Park" in cell.value:
            cell.offset(0,1).value = "Hollywood Park"
        elif "Live Oak" in cell.value:
            cell.offset(0,1).value = "Live Oak"
        elif "Converse" in cell.value:
            cell.offset(0,1).value = "Converse"

current_time = time.strftime("%Y_%m_%d%H_%M_%S")
# print(current_time)
workbook.save('addresses_cities{}.xlsx'.format(current_time))
