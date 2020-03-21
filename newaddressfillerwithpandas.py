import openpyxl
import os
import glob
import pandas
from openpyxl import load_workbook
from openpyxl import Workbook
import selenium
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options 
import tkinter
from tkinter import filedialog

driver = webdriver.Firefox()

donortestinfo = pandas.read_excel('testsheet.xlsx')
donors = donortestinfo.to_dict('records')


names = []
addresses = []
city = []

for record in donors:
    driver.get('http://www.bcad.org/clientdb/propertysearch.aspx?cid=1')
    search_term = donors["Name"]
    try:
            driver.find_element_by_id('propertySearchOptions_searchText').send_keys(search_term)
            driver.find_element_by_id('propertySearchOptions_search').click()
            possible_address = driver.find_element_by_id('propertySearchResults_address0')
    except selenium.common.exceptions.NoSuchElementException:
            driver.get('https://www.truepeoplesearch.com/')
            driver.find_element_by_id("Name").send_keys(search_term)
            driver.find_element_by_id("CityStateZip").send_keys(str(donors["City"]))
            driver.find_element_by_id("btnSubmit").click()
            viewAllDetailsButtons = driver.find_elements_by_link_text("View All Details")
            viewAllDetailsButtons[0].click()
            addressResult = driver.find_element_by_class_name("link-to-more")
            addresses.append(addressResult.text)
    else:
            addresses.append(possible_address.text)

driver.quit()

print(addresses)


# for name in names:
#     driver.get('http://www.bcad.org/clientdb/propertysearch.aspx?cid=1')
#     search_term = str(name)
#     try:
#         driver.find_element_by_id('propertySearchOptions_searchText').send_keys(search_term)
#         driver.find_element_by_id('propertySearchOptions_search').click()
#         possible_address = driver.find_element_by_id('propertySearchResults_address0')
#     except selenium.common.exceptions.NoSuchElementException:
#             driver.get('https://www.truepeoplesearch.com/')
#             driver.find_element_by_id("Name").send_keys(search_term)
#             driver.find_element_by_id("CityStateZip").send_keys("San Antonio, TX")
#             driver.find_element_by_id("btnSubmit").click()
#             viewAllDetailsButtons = driver.find_elements_by_link_text("View All Details")
#             viewAllDetailsButtons[0].click()
#             addressResult = driver.find_element_by_class_name("link-to-more")
#             addresses.append(addressResult.text)
#         # else:
#         #     print('failed to get address')
#         #     addresses.append('No Address Found')
#     # continue
    
#     else:
#         addresses.append(possible_address.text)

# driver.quit()

# print(addresses)
# col = 2
# start_row = 1
# for address in addresses:
#     ws1.cell(row=start_row, column=2).value= str(address).title()
#     start_row += 1

# addressesUpdatedSheetName = tkinter.filedialog.asksaveasfilename()
# workbook.save(addressesUpdatedSheetName)
# workbook.save('addresses_updated.xlsx')