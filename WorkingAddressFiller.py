import openpyxl
import pandas
from openpyxl import load_workbook
from openpyxl import Workbook
import selenium
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
import tkinter
from tkinter import filedialog

driver = webdriver.Firefox()
addressSearchSheet = tkinter.filedialog.askopenfilename()

workbook = load_workbook(addressSearchSheet)
sheets = workbook.sheetnames
ws1 = workbook[sheets[0]]
names = []
addresses = []

for row in ws1.iter_rows(min_row=1, max_row=None, min_col=1, max_col=1):
    for cell in row:
        names.append(cell.value)

for name in names:
    driver.get('http://www.bcad.org/clientdb/propertysearch.aspx?cid=1')
    search_term = str(name)
    try:
        driver.find_element_by_id(
            'propertySearchOptions_searchText').send_keys(search_term)
        driver.find_element_by_id('propertySearchOptions_search').click()
        possible_address = driver.find_element_by_id(
            'propertySearchResults_address0')
    except selenium.common.exceptions.NoSuchElementException:
        # Function code block which checks TruePeopleSearch.com when no result is found. This is omitted due to issue with Captchas
        # driver.get('https://www.truepeoplesearch.com/')
        # driver.find_element_by_id("Name").send_keys(search_term)
        # driver.find_element_by_id("CityStateZip").send_keys("San Antonio, TX")
        # driver.find_element_by_id("btnSubmit").click()
        # viewAllDetailsButtons = driver.find_elements_by_link_text("View All Details")
        # viewAllDetailsButtons[0].click()
        # addressResult = driver.find_element_by_class_name("link-to-more")
        # addresses.append(addressResult.text)
        # else:
        print('failed to get address')
        addresses.append('No Address Found')
    # continue

    else:
        addresses.append(possible_address.text)

driver.quit()

col = 2
start_row = 1

# insert returned address next to existing name in Excel worksheet
for address in addresses:
    ws1.cell(row=start_row, column=2).value = str(address).title()
    start_row += 1

addressesUpdatedSheetName = tkinter.filedialog.asksaveasfilename()
workbook.save(addressesUpdatedSheetName)
