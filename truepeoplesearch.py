import openpyxl
import os
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
import selenium
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options 

# This is a function code block reading the first column of the excel sheet and creating a list of values.
workbook = load_workbook('mailing_list.xlsx')
sheets = workbook.sheetnames
ws1 = workbook[sheets[0]]
driver = webdriver.Firefox()

names = []
addresses = []
for row in ws1.iter_rows(min_row=1,max_row=None, min_col=1, max_col =1):
    for cell in row:
        names.append(cell.value)
        
print(names)


for name in names:
    search_term = str(name)

        # this code block works and checks true people search but 
    driver.get('https://www.truepeoplesearch.com/')
    driver.find_element_by_id("Name").send_keys(search_term)
    driver.find_element_by_id("CityStateZip").send_keys("San Antonio, TX")
    driver.find_element_by_id("btnSubmit").click()
    viewAllDetailsButtons = driver.find_elements_by_link_text("View All Details")
    viewAllDetailsButtons[0].click()
    addressResult = driver.find_element_by_class_name("link-to-more")
    addresses.append(addressResult.text)

driver.quit()

col = 2
start_row = 1
for address in addresses:
    ws1.cell(row=start_row, column=2).value= str(address).title()
    start_row += 1

addressesUpdatedSheetName = tkinter.filedialog.asksaveasfilename()
workbook.save(addressesUpdatedSheetName)

print(addressResult.text)
driver.quit()